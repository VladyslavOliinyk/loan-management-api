from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from sqlalchemy import and_, func
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Credit, Dictionary, Payment, Plan
from app.schemas import PlanPerformanceItem, PlanPerformanceResponse, PlansInsertResponse

router = APIRouter()


@router.post("/plans_insert", response_model=PlansInsertResponse)
def insert_plans(file: UploadFile, db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx)")

    wb = load_workbook(file.file, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    categories = {d.name: d.id for d in db.query(Dictionary).all()}
    errors = []
    plans_to_insert = []

    for i, row in enumerate(rows, start=2):
        if len(row) < 3:
            errors.append(f"Row {i}: not enough columns")
            continue

        period_val, category_name, sum_val = row[0], row[1], row[2]

        if period_val is None:
            errors.append(f"Row {i}: period is empty")
            continue

        if isinstance(period_val, str):
            try:
                from datetime import datetime
                period_val = datetime.strptime(period_val, "%d.%m.%Y").date()
            except ValueError:
                errors.append(f"Row {i}: invalid date format '{period_val}', expected DD.MM.YYYY")
                continue

        if hasattr(period_val, "date"):
            period_val = period_val.date()

        if period_val.day != 1:
            errors.append(f"Row {i}: period must be the first day of the month, got {period_val}")
            continue

        if sum_val is None:
            errors.append(f"Row {i}: sum is empty")
            continue

        if category_name not in categories:
            errors.append(f"Row {i}: unknown category '{category_name}'")
            continue

        category_id = categories[category_name]

        existing = (
            db.query(Plan)
            .filter(Plan.period == period_val, Plan.category_id == category_id)
            .first()
        )
        if existing:
            errors.append(
                f"Row {i}: plan for {period_val} / {category_name} already exists"
            )
            continue

        plans_to_insert.append(
            Plan(period=period_val, sum=float(sum_val), category_id=category_id)
        )

    if errors:
        raise HTTPException(status_code=400, detail=errors)

    db.add_all(plans_to_insert)
    db.commit()

    return PlansInsertResponse(message=f"Successfully inserted {len(plans_to_insert)} plans")


@router.get("/plans_performance", response_model=PlanPerformanceResponse)
def get_plans_performance(check_date: date = Query(...), db: Session = Depends(get_db)):
    plans = (
        db.query(Plan, Dictionary.name)
        .join(Dictionary, Plan.category_id == Dictionary.id)
        .filter(Plan.period <= check_date)
        .all()
    )

    items = []
    for plan, category_name in plans:
        period_start = plan.period
        period_end = check_date

        if category_name == "видача":
            fact = (
                db.query(func.coalesce(func.sum(Credit.body), 0))
                .filter(
                    and_(
                        Credit.issuance_date >= period_start,
                        Credit.issuance_date <= period_end,
                    )
                )
                .scalar()
            )
        elif category_name == "збір":
            fact = (
                db.query(func.coalesce(func.sum(Payment.sum), 0))
                .filter(
                    and_(
                        Payment.payment_date >= period_start,
                        Payment.payment_date <= period_end,
                    )
                )
                .scalar()
            )
        else:
            continue

        fact = float(fact)
        plan_sum = float(plan.sum)
        perf = (fact / plan_sum * 100) if plan_sum else 0

        items.append(
            PlanPerformanceItem(
                month=plan.period,
                category=category_name,
                plan_sum=plan.sum,
                fact_sum=fact,
                performance_percent=round(perf, 2),
            )
        )

    return PlanPerformanceResponse(items=items)
